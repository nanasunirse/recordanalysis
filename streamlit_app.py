import streamlit as st
from pathlib import Path

logo_path = Path("genelater_logo.png")

if logo_path.exists():
    st.image(str(logo_path), width=700)
else:
    st.warning("Logo file not found")
    
# Generated from: Perfect_6.25.ipynb
# Converted at: 2026-08-01T10:35:39.795Z
# Next step (optional): refactor into modules & generate tests with RunCell
# Quick start: pip install runcell

# -*- coding: utf-8 -*-
from pathlib import Path
import re, json, math, html, base64, io
from collections import Counter
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

DEFAULT_BASE = Path.cwd()
FALLBACK_BASE = Path('/mnt/data')

def resolve_base(base_dir=None):
    if base_dir is not None:
        return Path(base_dir)
    cwd = Path.cwd()
    if list(cwd.glob('Report_*.xlsm')):
        return cwd
    if list(FALLBACK_BASE.glob('Report_*.xlsm')):
        return FALLBACK_BASE
    return cwd

ICONS = {
    'book': '\U0001F4D8', 'date': '\U0001F4C5', 'trophy': '\U0001F3C6', 'fire': '\U0001F525',
    'memo': '\U0001F4DD', 'pie': '\U0001F967', 'chart': '\U0001F4C8', 'target': '\U0001F3AF',
    'lab': '\U0001F9EA', 'sweet': '\U0001F9CB', 'alcohol': '\U0001F377', 'calories': '\U0001F525',
    'fat': '\U0001F951', 'carbs': '\U0001F359', 'protein': '\U0001F356', 'satfat': '\U0001F9C8',
    'transfat': '\U0001F35F', 'fiber': '\U0001F966', 'sodium': '\U0001F9C2', 'calcium': '\U0001F9B4',
    'ok': '\u2705'
}

DAY_RE = re.compile(r'^[A-Za-z]+,\s+[A-Za-z]+\s+\d{1,2}$')
MEALS = {'Breakfast', 'Lunch', 'Dinner', 'Snacks'}
ADEQUACY = {'Total Fat', 'Carbohydrates', 'Protein', 'Saturated Fat', 'Fiber', 'Calcium'}
MODERATION = {'Trans Fat', 'Sodium'}
DISPLAY = {
    'calories': 'Calories', 'fat_g': 'Total Fat', 'carbs_g': 'Carbohydrates', 'protein_g': 'Protein',
    'sat_fat_g': 'Saturated Fat', 'trans_fat_g': 'Trans Fat', 'fiber_g': 'Fiber', 'sodium_mg': 'Sodium',
    'calcium_mg': 'Calcium'
}
NUTRIENT_ORDER = ['fat_g', 'carbs_g', 'protein_g', 'sat_fat_g', 'trans_fat_g', 'fiber_g', 'sodium_mg', 'calcium_mg']
TARGET_ALIASES = {
    'calories': ['calories'], 'fat_g': ['total fat'], 'carbs_g': ['total carbs', 'carbs', 'carbohydrates'],
    'protein_g': ['protein'], 'sat_fat_g': ['sat fat', 'saturated fat'], 'trans_fat_g': ['trans fat'],
    'fiber_g': ['fiber'], 'sodium_mg': ['sodium'], 'calcium_mg': ['calcium']
}
TARGET_DEFAULTS = {'calories': 2000.0, 'fat_g': 70.0, 'carbs_g': 250.0, 'protein_g': 90.0, 'sat_fat_g': 20.0, 'trans_fat_g': 0.0, 'fiber_g': 20.0, 'sodium_mg': 2300.0, 'calcium_mg': 1000.0}
NUTRIENT_ICON_KEY = {'Calories': 'calories', 'Total Fat': 'fat', 'Carbohydrates': 'carbs', 'Protein': 'protein', 'Saturated Fat': 'satfat', 'Trans Fat': 'transfat', 'Fiber': 'fiber', 'Sodium': 'sodium', 'Calcium': 'calcium'}

ADD_FOODS = {
    'Calories': ["\U0001F35A Rice (1 bowl per meal)", "\U0001F35E Bread (2 slices)", "\U0001F34C Banana (1 piece)"],
    'Total Fat': ["\U0001F951 Avocado (1/2 fruit per meal)", "\U0001F95C Mixed nuts (30g per day)", "\U0001FAD2 Olive oil (1 tbsp per meal)"],
    'Carbohydrates': ["\U0001F35A Rice (1 bowl per meal)", "\U0001F360 Sweet potato (100g per meal)", "\U0001F35E Bread (2 slices)"],
    'Protein': ["\U0001F95A Egg (1 egg per meal)", "\U0001F41F Chicken or fish (100g per meal)", "\U0001F9CA Tofu or yogurt (150g tofu / 1 cup yogurt per day)"],
    'Saturated Fat': ["\U0001F95B Milk (1 cup per day)", "\U0001F9C0 Cheese (30g per day)", "\U0001F95A Egg (1 egg per meal)"],
    'Fiber': ["\U0001F966 Vegetables (1 cup per meal)", "\U0001F34E Fruit (1 piece per day)", "\U0001FAD8 Beans or lentils (1/2 cup per day)"],
    'Calcium': ["\U0001F95B Milk (1 cup per day)", "\U0001F963 Yogurt (1 cup per day)", "\U0001F9CA Tofu (150g per meal)"]
}
REDUCE_FOODS = {
    'Total Fat': ["\U0001F357 Fried foods: reduce 1 serving per day", "\U0001F96B Creamy sauces: use less", "\U0001F969 Fatty meat: choose lean meat instead"],
    'Carbohydrates': ["\U0001F35A Rice/noodles: reduce about 1/2 bowl per meal", "\U0001F370 Sweets: reduce 1 serving per day", "\U0001F9CB Sugary drinks: reduce 1 bottle/can per day"],
    'Protein': ["\U0001F969 Large meat portions: keep to about 100g per meal", "\U0001F357 Avoid several big meat dishes in one day", "\U0001F953 Limit processed meat like bacon or sausage"],
    'Saturated Fat': ["\U0001F9C8 Butter/cream: use less", "\U0001F950 Pastries: reduce 1 serving per day", "\U0001F357 Choose grilled dishes more often"],
    'Trans Fat': ["\U0001F35F Packaged fried snacks: reduce 1 serving per day", "\U0001F950 Fast food/pastries: reduce 1 serving per day", "\U0001F373 Choose simple cooked foods more often"],
    'Sodium': ["\U0001F35C Soup/instant noodles: reduce 1 bowl/pack per day", "\U0001F9C2 Soy sauce/fish sauce: use half the usual amount", "\U0001F952 Pickled/processed foods: reduce 1 serving per day"]
}
SWEET_PATTERNS = ['bubble tea', 'milk tea', 'latte', 'cake', 'dessert', 'cookie', 'ice cream', 'coke', 'cola', 'soft drink', 'syrup', 'sweetened', 'iced coffee with milk', 'shaved ice dessert', 'fruit dessert', 'croissant', 'pastry']
ALCOHOL_WORDS = ['wine', 'beer', 'cabernet', 'whisky', 'whiskey', 'vodka', 'rum', 'gin', 'tequila', 'sake', 'soju', 'cider']
ALCOHOL_RE = re.compile(r'\b(?:' + '|'.join(sorted(map(re.escape, ALCOHOL_WORDS), key=len, reverse=True)) + r')\b', re.I)

CSS = """
@page { size:A4; margin:10.5mm; }
body { font-family:'Arial','Noto Color Emoji','Segoe UI Emoji','Apple Color Emoji',sans-serif; margin:0; color:#1f2937; background:#f5f7fb; font-size:10pt; line-height:1.22; }
h1,h2,h3 { margin:0 0 5px 0; } h1 { font-size:22pt; color:#0f2b46; } h2 { font-size:12.4pt; color:#0f2b46; } h3 { font-size:10.6pt; color:#20486d; }
.sub { color:#667085; margin-bottom:8px; font-size:9pt; }
.grid { display:grid; grid-template-columns:repeat(3,1fr); gap:8px; margin-bottom:8px; }
.card,.section { background:#fff; border:1px solid #d7dde7; border-radius:12px; padding:9px; box-shadow:0 1px 2px rgba(17,24,39,.05); break-inside:avoid; page-break-inside:avoid; }
.metric { font-size:22px; font-weight:700; margin-top:4px; color:#0f2b46; }
.two-col { display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:8px; }
.stack { display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:8px; }
ul { margin:6px 0 0 16px; padding-left:0; } li { margin-bottom:4px; }
img { max-width:100%; border:1px solid #e8edf5; border-radius:8px; }
table.data-table { width:100%; border-collapse:collapse; font-size:8.7pt; }
table.data-table th, table.data-table td { padding:4px 5px; border-bottom:1px solid #edf2f7; text-align:left; vertical-align:top; }
table.data-table th { background:#f8fafc; color:#334155; }
.small { color:#667085; font-size:8pt; margin-bottom:4px; }
.badge { display:inline-block; padding:2px 7px; border-radius:999px; font-weight:700; font-size:8pt; }
.good { background:#E7F6EA; color:#2E7D32; } .warn { background:#FFF7E0; color:#A66A00; } .bad { background:#FDECEC; color:#C62828; }
.callout { border-left:5px solid #2563eb; background:#eff6ff; padding:8px 10px; border-radius:9px; margin:7px 0 8px 0; font-size:8.8pt; }
.ai-badge { display:inline-block; padding:2px 8px; border-radius:999px; background:#eef2ff; color:#3730a3; font-size:8pt; font-weight:700; }
"""

def strip_html(s):
    s = '' if s is None else str(s)
    s = re.sub(r'<[^>]+>', '', s)
    return s.replace('&nbsp;', ' ').replace('\xa0', ' ').strip()

def clean(s): return re.sub(r'\s+', ' ', strip_html(s)).strip()
def normalize(s): return re.sub(r'[^a-z0-9]+', ' ', clean(s).lower()).strip()
def esc(s): return html.escape(str(s))
def n(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = clean(v)
    if not s or s in {'g', 'mg', 'cals'}: return None
    m = re.search(r'[-+]?\d[\d,]*\.?\d*', s)
    return float(m.group().replace(',', '')) if m else None

def grade(v):
    m = re.search(r'([ABCDF][+-]?)', clean(v))
    return m.group(1) if m else clean(v)

def unit_for(key): return 'mg' if key.endswith('_mg') else ('kcal' if key == 'calories' else 'g')
def fmt(v, unit):
    if v is None or (isinstance(v, float) and math.isnan(v)): return '-'
    v = float(v)
    return f'{v:.0f} {unit}' if unit in ('mg','kcal') or abs(v-round(v)) < 1e-9 else f'{v:.1f} {unit}'
def nutrient_label(name): return ICONS[NUTRIENT_ICON_KEY[name]] + name

def food_label(name):
    raw = clean(name)
    low = raw.lower()
    rules = [
        (['pizza'], '\U0001F355'), (['rice porridge','porridge','congee','rice'], '\U0001F35A'),
        (['udon','spagehetti','spaghetti','noodle','rice noodle','pasta'], '\U0001F35C'),
        (['duck','pork','beef','steak','chicken','lamb','rib','patty','sirloin','salami'], '\U0001F969'),
        (['fish','shrimp','octopus','salmon','seafood'], '\U0001F41F'), (['sandwich','bread','croissant','cracker'], '\U0001F35E'),
        (['milk','soy milk','yogurt','latte'], '\U0001F95B'), (['cake','dessert','pudding','cookie','ice'], '\U0001F37D'),
        (['coffee','tea'], '\u2615'), (['mango','orange','banana','watermelon','fruit','plum','papaya','melon','durian'], '\U0001F34E'),
        (['vegetable','spinach','cabbage','lettuce','cucumber','carrot','eggplant','greens','bean sprout'], '\U0001F957'),
        (['egg','omelette'], '\U0001F95A'), (['tofu'], '\U0001F9CA'), (['soup'], '\U0001F372'), (['dumpling'], '\U0001F95F')]
    for needles, emo in rules:
        if any(x in low for x in needles): return emo + ' ' + esc(raw)
    if ALCOHOL_RE.search(low): return '\U0001F377 ' + esc(raw)
    return '\U0001F37D ' + esc(raw)

def open_rows(path):
    wb = load_workbook(path, data_only=True, keep_vba=True)
    ws = wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]

def detect_header_row(rows):
    for i, row in enumerate(rows[:30]):
        cells = [normalize(c) for c in row if clean(c)]
        score = sum([any(c == 'calories' for c in cells), any(c.startswith('total fat') for c in cells), any(c.startswith('carbs') or c.startswith('total carbs') or c.startswith('carbohydrates') for c in cells), any(c.startswith('protein') for c in cells), any('sodium' in c for c in cells), any('calcium' in c for c in cells)])
        if score >= 5: return i
    raise ValueError('Header row not found')

def classify_header(cell):
    t = normalize(cell)
    if not t: return None
    if 'fd grade' in t or t == 'grade': return 'food_grade'
    if t == 'calories' or ' calories' in t: return 'calories'
    if 'trans fat' in t: return 'trans_fat_g'
    if 'sat fat' in t or 'saturated fat' in t: return 'sat_fat_g'
    if 'total fat' in t: return 'fat_g'
    if 'total carbs' in t or 'carbohydrate' in t or t.startswith('carbs'): return 'carbs_g'
    if 'protein' in t: return 'protein_g'
    if 'fiber' in t: return 'fiber_g'
    if 'sodium' in t: return 'sodium_mg'
    if 'calcium' in t: return 'calcium_mg'
    return None

def build_colmap(header_row):
    out = {}
    for idx, cell in enumerate(header_row):
        key = classify_header(cell)
        if key and key not in out: out[key] = idx
    return out

def parse_targets(rows, header_idx):
    targets = {}
    for i in range(min(12, len(rows))):
        cells = [clean(x) for x in rows[i] if clean(x)]
        if any(c.lower() == 'target' for c in cells):
            for j in range(i+1, min(header_idx, i+10)):
                vals = [clean(x) for x in rows[j] if clean(x)]
                for cell in vals:
                    m = re.match(r'^([\d,]+(?:\.\d+)?)\s*,\s*(.+)$', cell)
                    if m:
                        num = float(m.group(1).replace(',', ''))
                        label = normalize(m.group(2))
                        for k, aliases in TARGET_ALIASES.items():
                            if any(a in label for a in aliases): targets[k] = num
            break
    for k, v in TARGET_DEFAULTS.items(): targets.setdefault(k, v)
    return targets

def row_metrics(row, cmap):
    d = {}
    for k, idx in cmap.items():
        val = row[idx] if idx < len(row) else None
        d[k] = grade(val) if k == 'food_grade' else n(val)
    return d

def parse_period(rows):
    title = clean(rows[0][0] if rows and rows[0] else '')
    m = re.search(r'for\s+([\d./-]+)\s+-\s+([\d./-]+)', title)
    return f'{m.group(1)} - {m.group(2)}' if m else title

def parse_workbook(path):
    rows = open_rows(path)
    header_idx = detect_header_row(rows)
    cmap = build_colmap(rows[header_idx])
    targets = parse_targets(rows, header_idx)
    period = parse_period(rows)
    avg_idx = None
    for i in range(header_idx+1, min(header_idx+8, len(rows))):
        if clean(rows[i][0] if rows[i] else '').startswith('Averages over the period'):
            avg_idx = i
            break
    if avg_idx is None: avg_idx = header_idx + 1
    avg = row_metrics(rows[avg_idx], cmap)
    days, foods = [], []
    current_day, current_meal = None, None
    for i in range(avg_idx+1, len(rows)):
        row = rows[i]; first = clean(row[0] if row else None)
        if not first: continue
        if DAY_RE.match(first):
            current_day = first; current_meal = None; rec = row_metrics(row, cmap); rec['date'] = current_day; days.append(rec); continue
        if first in MEALS: current_meal = first; continue
        if first == 'Left': continue
        if current_day:
            rec = row_metrics(row, cmap)
            if any(rec.get(k) is not None for k in ['calories','fat_g','carbs_g','protein_g','sat_fat_g','fiber_g','sodium_mg','calcium_mg']):
                rec['date'] = current_day; rec['meal'] = current_meal or ''; rec['food_name'] = first; foods.append(rec)
    return {'source': path.name, 'period': period, 'header_row': header_idx+1, 'column_map': cmap, 'targets': targets, 'avg': avg, 'days': pd.DataFrame(days), 'foods': pd.DataFrame(foods)}

def classify_nutrients(avg, targets):
    out = []
    for key in NUTRIENT_ORDER:
        intake, target = avg.get(key), targets.get(key)
        if intake is None or target is None: continue
        name = DISPLAY[key]; pct = 0.0 if not target else float(intake)/float(target)*100
        if name in MODERATION:
            if key == 'trans_fat_g' and target == 0:
                status = 'Good' if float(intake) == 0 else 'Too much'; pct = 0.0 if float(intake) == 0 else 999.0
            else: status = 'Too much' if pct > 120 else 'Good'
        else:
            if pct < 85: status = 'Not enough'
            elif pct > 120: status = 'Too much'
            else: status = 'Good'
        out.append({'key': key, 'name': name, 'intake': float(intake), 'target': float(target), 'pct': pct, 'status': status, 'unit': unit_for(key)})
    return out

def overall_score(classified):
    vals = []
    for r in classified:
        pct = r['pct']
        if r['name'] in MODERATION: s = 10 if r['status'] == 'Good' else max(0, 10 - max(0, pct - 100) / 10)
        else: s = max(0, min(100, pct) / 10) if pct <= 100 else max(0, 10 - (pct - 100) / 10)
        vals.append(s)
    return int(round(sum(vals)/len(vals)*10)) if vals else 0

def pfc(avg):
    p = (avg.get('protein_g') or 0) * 4; f = (avg.get('fat_g') or 0) * 9; c = (avg.get('carbs_g') or 0) * 4; total = p + f + c
    return {'Protein': 0.0, 'Fat': 0.0, 'Carbohydrates': 0.0} if not total else {'Protein': round(p/total*100,1), 'Fat': round(f/total*100,1), 'Carbohydrates': round(c/total*100,1)}

def top_foods(food_df, key, n_top=3):
    if food_df.empty or key not in food_df.columns: return []
    tmp = food_df[['food_name', key]].copy(); tmp[key] = pd.to_numeric(tmp[key], errors='coerce'); tmp = tmp.dropna(subset=[key]); tmp = tmp[tmp[key] > 0]; tmp = tmp.sort_values(key, ascending=False).drop_duplicates('food_name').head(n_top)
    return tmp.to_dict('records')

def build_add_reduce(classified, foods_df):
    add_rows, reduce_rows = [], []
    for r in classified:
        name, key, unit = r['name'], r['key'], r['unit']
        if name in ADEQUACY and r['status'] == 'Not enough':
            gap = max(0, r['target'] - r['intake']); tops = top_foods(foods_df, key, 3); sugg = ADD_FOODS.get(name, ['-','-','-'])
            for i in range(3):
                top = f"{food_label(tops[i]['food_name'])}: {fmt(tops[i][key], unit)}" if i < len(tops) else ''
                add_rows.append({'Nutrient': nutrient_label(name) if i == 0 else '', '\U0001F534 Current Gap': f"-{fmt(gap, unit)} ({r['pct']:.0f}%)" if i == 0 else '', '\U0001F4CA Your Top 3 Food (from data)': top, 'Suggested food': sugg[i] if i < len(sugg) else ''})
        elif name in ADEQUACY and r['status'] == 'Too much':
            excess = max(0, r['intake'] - r['target']); tops = top_foods(foods_df, key, 3); sugg = REDUCE_FOODS.get(name, ['-','-','-'])
            for i in range(3):
                top = f"{food_label(tops[i]['food_name'])}: {fmt(tops[i][key], unit)}" if i < len(tops) else ''
                reduce_rows.append({'Nutrient': nutrient_label(name) if i == 0 else '', '\U0001F534 Current Excess': f"+{fmt(excess, unit)} ({r['pct']:.0f}%)" if i == 0 else '', '\U0001F4CA Your Top 3 Food (from data)': top, 'Suggested food': sugg[i] if i < len(sugg) else ''})
        elif name in MODERATION and r['status'] == 'Too much':
            excess = r['intake'] if key == 'trans_fat_g' and r['target'] == 0 else max(0, r['intake'] - r['target']); tops = top_foods(foods_df, key, 3); sugg = REDUCE_FOODS.get(name, ['-','-','-'])
            for i in range(3):
                top = f"{food_label(tops[i]['food_name'])}: {fmt(tops[i][key], unit)}" if i < len(tops) else ''
                reduce_rows.append({'Nutrient': nutrient_label(name) if i == 0 else '', '\U0001F534 Current Excess': f"+{fmt(excess, unit)} ({r['pct']:.0f}%)" if i == 0 else '', '\U0001F4CA Your Top 3 Food (from data)': top, 'Suggested food': sugg[i] if i < len(sugg) else ''})
    return pd.DataFrame(add_rows), pd.DataFrame(reduce_rows)

def meal_logging(days_df, foods_df, total_days=30):
    logged_days = int(days_df.shape[0]) if not days_df.empty else 0
    meal_sets = {m: set(foods_df.loc[foods_df['meal'] == m, 'date']) if not foods_df.empty else set() for m in ['Breakfast','Lunch','Dinner']}
    rates = {m: len(v)/total_days for m, v in meal_sets.items()}
    df = pd.DataFrame([
        {'Meal': 'Total logged days', 'Recorded days': logged_days, 'Out of 30 days': f'{logged_days}/30 ({logged_days/30*100:.1f}%)'},
        {'Meal': 'Breakfast', 'Recorded days': len(meal_sets['Breakfast']), 'Out of 30 days': f"{len(meal_sets['Breakfast'])}/30 ({rates['Breakfast']*100:.1f}%)"},
        {'Meal': 'Lunch', 'Recorded days': len(meal_sets['Lunch']), 'Out of 30 days': f"{len(meal_sets['Lunch'])}/30 ({rates['Lunch']*100:.1f}%)"},
        {'Meal': 'Dinner', 'Recorded days': len(meal_sets['Dinner']), 'Out of 30 days': f"{len(meal_sets['Dinner'])}/30 ({rates['Dinner']*100:.1f}%)"},
    ])
    return rates, df

def sweet_alcohol(foods_df):
    if foods_df.empty: return 0, pd.DataFrame(columns=['Top 3 drinks/snacks','Count']), 0, pd.DataFrame(columns=['Top 3 alcohol drinks','Count'])
    names = foods_df['food_name'].fillna('').map(clean)
    sweet_mask = names.str.lower().apply(lambda s: any(p in s for p in SWEET_PATTERNS))
    alc_mask = names.str.lower().apply(lambda s: bool(ALCOHOL_RE.search(s)))
    sweet_df, alc_df = foods_df[sweet_mask].copy(), foods_df[alc_mask].copy()
    sweet_count, alc_count = int(len(sweet_df)), int(len(alc_df))
    sweet_freq = pd.DataFrame(Counter([food_label(x) for x in sweet_df['food_name']]).most_common(3), columns=['Top 3 drinks/snacks','Count']) if sweet_count else pd.DataFrame(columns=['Top 3 drinks/snacks','Count'])
    alc_freq = pd.DataFrame(Counter([food_label(x) for x in alc_df['food_name']]).most_common(3), columns=['Top 3 alcohol drinks','Count']) if alc_count else pd.DataFrame(columns=['Top 3 alcohol drinks','Count'])
    return sweet_count, sweet_freq, alc_count, alc_freq

def daily_score(row, targets):
    vals = []
    for key in NUTRIENT_ORDER:
        intake, target = row.get(key), targets.get(key)
        if intake is None or pd.isna(intake) or target is None: continue
        pct = 0 if target == 0 else float(intake)/float(target)*100
        if DISPLAY[key] in MODERATION: s = 10 if pct <= 100 else max(0, 10 - (pct - 100) / 10)
        else: s = max(0, min(pct, 100) / 10) if pct <= 100 else max(0, 10 - (pct - 100) / 10)
        vals.append(s)
    return round(sum(vals)/len(vals)*10, 1) if vals else None

def best_worst(days_df, targets):
    if days_df.empty: return pd.DataFrame(columns=['Section','Date','Score','Calories','Comment'])
    x = days_df.copy(); x['daily_score'] = x.apply(lambda r: daily_score(r, targets), axis=1); x['calories_num'] = pd.to_numeric(x['calories'], errors='coerce'); x = x.dropna(subset=['daily_score'])
    if x.empty: return pd.DataFrame(columns=['Section','Date','Score','Calories','Comment'])
    best = x.sort_values(['daily_score','calories_num'], ascending=[False,True]).iloc[0]; worst = x.sort_values(['daily_score','calories_num'], ascending=[True,False]).iloc[0]
    return pd.DataFrame([{'Section':'\u2B50 The best nutrition day','Date':best['date'],'Score':f"{float(best['daily_score']):.1f}",'Calories':f"{float(best['calories_num']):.0f} kcal",'Comment':'Closer to the goal and more balanced overall.'},{'Section':'\u26A0\uFE0F The worst nutrition day','Date':worst['date'],'Score':f"{float(worst['daily_score']):.1f}",'Calories':f"{float(worst['calories_num']):.0f} kcal",'Comment':'Farther from the goal and less balanced overall.'}])

def detailed_analysis(classified):
    rows = []
    for r in classified:
        badge = '<span class="badge good">good</span>' if r['status'] == 'Good' else ('<span class="badge warn">not enough</span>' if r['status'] == 'Not enough' else '<span class="badge bad">too much</span>')
        rows.append({'Nutrient': nutrient_label(r['name']), 'Intake': fmt(r['intake'], r['unit']), 'Target': fmt(r['target'], r['unit']), '% of target': f"{r['pct']:.0f}%", 'Status': badge})
    return pd.DataFrame(rows)

def local_actions(pfc_data, add_df, reduce_df, meal_rates, sweet_count, alcohol_count):
    actions = []
    p, f, c = pfc_data['Protein'], pfc_data['Fat'], pfc_data['Carbohydrates']
    if 15 <= p <= 25 and 20 <= f <= 35 and 45 <= c <= 65: actions.append('PFC balance: Your protein, fat, and carbohydrate balance is fairly close to a good pattern.')
    else:
        if p < 15: actions.append('PFC balance: Protein looks low. Try adding eggs, fish, chicken, tofu, or yogurt.')
        elif f > 35: actions.append('PFC balance: Fat looks high. Try using less fried food and creamy food.')
        elif c > 65: actions.append('PFC balance: Carbohydrates look high. Try reducing part of rice, noodles, bread, or sweets.')
        elif c < 45: actions.append('PFC balance: Carbohydrates look low. Try adding rice, bread, or sweet potato.')
        else: actions.append('PFC balance: The overall balance is a little uneven.')
    used = set()
    if not add_df.empty:
        for _, row in add_df.iterrows():
            nutrient, suggest = row.get('Nutrient',''), row.get('Suggested food','')
            if nutrient and suggest:
                nutrient_name = re.sub(r'^[^A-Za-z]+', '', re.sub(r'<[^>]+>', '', str(nutrient))).strip().lower()
                if nutrient_name not in used:
                    actions.append(f'Add more for {nutrient_name}: {re.sub(r"<[^>]+>", "", suggest)}'); used.add(nutrient_name)
                if len(used) >= 2: break
    if not reduce_df.empty:
        for _, row in reduce_df.iterrows():
            nutrient, suggest = row.get('Nutrient',''), row.get('Suggested food','')
            if nutrient and suggest:
                nutrient_name = re.sub(r'^[^A-Za-z]+', '', re.sub(r'<[^>]+>', '', str(nutrient))).strip().lower()
                actions.append(f'Reduce for {nutrient_name}: {re.sub(r"<[^>]+>", "", suggest)}'); break
    low_meals = [m.lower() for m, r in meal_rates.items() if r < 0.6]
    if low_meals: actions.append(f'Meal logging: {", ".join(low_meals)} was recorded less often. Recording it more regularly will make the report more accurate.')
    if sweet_count > 0: actions.append(f'Sweet drinks or sweets appeared {sweet_count} time(s). Reducing these first may help lower extra sugar and calories.')
    if alcohol_count > 0: actions.append(f'Alcohol appeared {alcohol_count} time(s). Keeping alcohol occasional may help overall balance.')
    return actions[:5]

def plot_to_data_uri(draw_fn):
    fig = plt.figure(figsize=(6,4), dpi=140); draw_fn(fig); buf = io.BytesIO(); plt.tight_layout(); plt.savefig(buf, format='png', dpi=140, bbox_inches='tight'); plt.close(fig)
    return 'data:image/png;base64,' + base64.b64encode(buf.getvalue()).decode('ascii')

def pfc_uri(pfc_data):
    def draw(fig):
        ax = fig.add_subplot(111); vals = list(pfc_data.values())
        if sum(vals) <= 0: ax.text(0.5,0.5,'No PFC data',ha='center',va='center'); ax.axis('off')
        else: ax.pie(vals, labels=list(pfc_data.keys()), autopct='%1.1f%%', startangle=90)
        ax.set_title('PFC Balance')
    return plot_to_data_uri(draw)

def daily_uri(days_df, target_cal):
    def draw(fig):
        ax = fig.add_subplot(111)
        if not days_df.empty:
            vals = pd.to_numeric(days_df['calories'], errors='coerce').fillna(0).tolist(); labels = [d.replace(',', '') for d in days_df['date'].astype(str).tolist()]
            ax.bar(range(len(vals)), vals, color='#5b8bd8'); ax.plot(range(len(vals)), [target_cal]*len(vals), linestyle='--', color='#d9534f')
            ax.set_xticks(range(len(vals))); ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=7); ax.set_ylabel('Calories'); ax.set_title('Daily calories')
        else: ax.text(0.5,0.5,'No daily data',ha='center',va='center'); ax.axis('off')
    return plot_to_data_uri(draw)

def nutrient_uri(classified):
    def draw(fig):
        ax = fig.add_subplot(111); labels = [r['name'] for r in classified]; perc = [r['pct'] for r in classified]; colors = ['#4CAF50' if r['status']=='Good' else ('#F4B400' if r['status']=='Not enough' else '#DB4437') for r in classified]; clipped = [min(v,200) for v in perc]
        ax.barh(range(len(labels)), clipped, color=colors); ax.axvline(100, linestyle='--', color='gray', linewidth=1); ax.set_xlim(0,220); ax.set_yticks(range(len(labels))); ax.set_yticklabels(labels); ax.set_xlabel('% of Target'); ax.set_title('Nutrient amount vs target')
        for y, p in enumerate(perc): ax.text(min(p+2,202), y, f'{p:.0f}%', va='center', ha='left', fontsize=8)
    return plot_to_data_uri(draw)

def table_html(df): return '<p>No data</p>' if df is None or df.empty else df.to_html(index=False, border=0, classes='data-table', escape=False)

def build_html(report):
    classified = classify_nutrients(report['avg'], report['targets']); score = overall_score(classified); pfc_data = pfc(report['avg']); add_df, reduce_df = build_add_reduce(classified, report['foods']); meal_rates, meal_df = meal_logging(report['days'], report['foods']); sweet_count, sweet_df, alc_count, alc_df = sweet_alcohol(report['foods']); detail_df = detailed_analysis(classified); best_worst_df = best_worst(report['days'], report['targets']); actions = local_actions(pfc_data, add_df, reduce_df, meal_rates, sweet_count, alc_count)
    action_items = ''.join(f'<li>{esc(x)}</li>' for x in actions) if actions else '<li>-</li>'
    return f'''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Nutrition Analysis Report</title><style>{CSS}</style></head><body>
    <h1>Nutrition Analysis Report</h1>
    <div class="sub">{ICONS['book']} <b>Source:</b> {esc(report['source'])} / {ICONS['date']} <b>Period:</b> {esc(report['period'])} / <span class="ai-badge">AI comment mode: local-fallback</span></div>
    <div class="grid"><div class="card"><div>{ICONS['trophy']} <b>Total nutrition score</b></div><div class="metric">{score}</div></div><div class="card"><div>{ICONS['fire']} <b>Average calories</b></div><div class="metric">{int(round(report['avg'].get('calories') or 0))}</div><div>Target {int(round(report['targets'].get('calories') or 0))}</div></div><div class="card"><div>{ICONS['memo']} <b>Variability</b></div><div class="metric">{int(report['days'].shape[0])}/30</div><div>logged days / 30 days</div></div></div>
    <div class="callout"><b>ℹ\uFE0F About Total nutrition score</b><br>This score is mimics The Healthy Eating Index (HEI) score, 🌟>80: Excellent! 👍 50-80: Good, but there's room for improvement. ⚠️< 50: Time to make some adjustments.  It shows how closely your eating habits match recommended guidelines for good health.</div>
    <div class="two-col"><div class="section"><h2>{ICONS['pie']} PFC pie chart</h2><div class="small">This shows how calories are split between protein, fat, and carbohydrates.</div><div class="small">This section only explains the overall balance of protein, fat, and carbohydrate.</div><img src="{pfc_uri(pfc_data)}"></div><div class="section"><h2>{ICONS['ok']} Recommended actions</h2><div class="small">Action points are generated only from PFC, add/reduce, meal logging, sweet drink frequency, and alcohol frequency.</div><ul>{action_items}</ul></div></div>
    <div class="section"><h2>{ICONS['date']} Meal recording over 30 days</h2><div class="small">This shows how often each meal was recorded over 30 days.</div>{table_html(meal_df)}</div>
    <div class="stack"><div class="section"><h2>{ICONS['chart']} Daily calories</h2><div class="small">This shows daily calories compared with the goal.</div><img src="{daily_uri(report['days'], report['targets'].get('calories',0))}"></div><div class="section"><h2>{ICONS['target']} Nutrient amount vs target</h2><div class="small">Green is okay, yellow is not enough, and red is too much. If a value is very high, the label still shows the full percentage.</div><img src="{nutrient_uri(classified)}"></div></div>
    <div class="two-col"><div class="section"><h2>{ICONS['lab']} Detailed nutrient analysis</h2><div class="small">This table shows amount-based classification for nutrients.</div>{table_html(detail_df)}</div><div class="section"><h2>\u2B50 Best day / \u26A0\uFE0F Worst day</h2><div class="small">These two days show the strongest and weakest overall balance.</div>{table_html(best_worst_df)}</div></div>
    <div class="section"><h2>\u2795 Foods to add</h2><div class="small">Only nutrients that are not enough are listed here.</div>{table_html(add_df)}</div>
    <div class="section"><h2>\u2796 Foods to reduce</h2><div class="small">Only nutrients that are too much are listed here.</div>{table_html(reduce_df)}</div>
    <div class="two-col"><div class="section"><h2>{ICONS['sweet']} Sweet drink frequency</h2><p><b>Count:</b> {sweet_count} record(s)</p>{table_html(sweet_df)}</div><div class="section"><h2>{ICONS['alcohol']} Alcohol frequency</h2><p><b>Count:</b> {alc_count} record(s)</p>{table_html(alc_df)}</div></div>
    </body></html>''', {'score': score, 'actions': actions, 'sweet_count': sweet_count, 'alcohol_count': alc_count, 'add_preview': add_df.head(4).to_dict('records'), 'reduce_preview': reduce_df.head(4).to_dict('records')}

def run_batch(input_files=None, base_dir=None):
    base = resolve_base(base_dir)
    files = [Path(p) for p in input_files] if input_files else sorted(base.glob('Report_*.xlsm'))
    if not files:
        raise FileNotFoundError(f'No Report_*.xlsm files found in {base}')
    manifest_path = base / 'nutrition_report_production_batch_notebook_manifest.json'
    manifest = []
    for path in files:
        if not path.is_absolute():
            path = base / path
        report = parse_workbook(path)
        html_text, meta = build_html(report)
        out = base / f'{path.stem}_ai_production_batch_final.html'
        out.write_text(html_text, encoding='utf-8')
        manifest.append({'source': report['source'], 'output_html': out.name, 'period': report['period'], 'header_row': report['header_row'], 'column_map': report['column_map'], **meta})
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return manifest

#if __name__ == '__main__':
#    run_batch()

# new input

import streamlit as st
import streamlit.components.v1 as components
import tempfile
from pathlib import Path

st.set_page_config(
    page_title="Nutrition Analysis Report",
    layout="wide"
)

st.title("Nutrition Analysis Report Generator")
st.write("Upload your food record Excel file, then the nutrition report will be generated automatically.")

uploaded_file = st.file_uploader(
    "Upload your Report Excel file",
    type=["xlsm", "xlsx"]
)

if uploaded_file is not None:
    st.success(f"Uploaded file: {uploaded_file.name}")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        # Save uploaded Excel file temporarily
        input_path = tmpdir / uploaded_file.name
        input_path.write_bytes(uploaded_file.getbuffer())

        try:
            # Run your existing main function
            manifest = run_batch(
                input_files=[input_path],
                base_dir=tmpdir
            )

            # Get output HTML file path
            output_html_name = manifest[0]["output_html"]
            output_html_path = tmpdir / output_html_name

            html_text = output_html_path.read_text(encoding="utf-8")

            st.subheader("Generated Nutrition Report")

            # Show HTML report inside Streamlit
            components.html(
                html_text,
                height=1200,
                scrolling=True
            )

            # Download button
            st.download_button(
                label="Download HTML Report",
                data=html_text,
                file_name=output_html_name,
                mime="text/html"
            )

        except Exception as e:
            st.error("An error occurred while generating the report.")
            st.exception(e)
else:
    st.info("Please upload an Excel file to start.")
