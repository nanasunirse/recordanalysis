#!/usr/bin/env python3
"""
Final MyNetDiary HTML Report Generator - Production 6-Chart Version

Purpose
- Read MyNetDiary Excel reports (.xlsx / .xlsm)
- Output separated A4-style HTML reports by person/file
- Fully functional all-in-one script; no manual modification required

Current logic preserved
- Include only days with >= 2 main meals recorded
- Main meals are Breakfast, Lunch, and Dinner. Snacks are not counted
- Use median and IQR for robust nutrient summaries
- Calories trend is shown in kcal, not percentage
- Other nutrition trend charts are shown as % of target
- Six nutrition trend charts are placed on the same A4 page
- Sodium and other extreme trend outliers are visually handled with an omitted wavy line
- Output format is HTML

Jupyter usage
    %run final_mynetdiary_html_report_6charts.py --files Report_043.xlsm Report_276.xlsm

Terminal usage
    python final_mynetdiary_html_report_6charts.py --files Report_043.xlsm Report_276.xlsm

If no --files are supplied, the script reads all .xlsx/.xlsm files in --input-dir.
"""

from __future__ import annotations

import argparse
import base64
import html
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================
MIN_MAIN_MEALS = 2
MAIN_MEALS = {"Breakfast", "Lunch", "Dinner"}

NUTRIENTS = [
    "Calories",
    "Total Fat, g",
    "Carbs, g",
    "Protein, g",
    "Sodium, mg",
    "Calcium, mg",
]

# 6 trend figures should be on the same A4 page.
# Calories is kcal. All others are % of target.
TREND_NUTRIENTS = [
    "Calories",
    "Total Fat, g",
    "Carbs, g",
    "Protein, g",
    "Sodium, mg",
    "Calcium, mg",
]

TARGET_KEYWORDS = {
    "Calories": "Calories",
    "Total Fat, g": "Total Fat",
    "Carbs, g": "Total Carbs",
    "Protein, g": "Protein",
    "Sodium, mg": "Sodium",
    "Calcium, mg": "Calcium",
}

UNITS = {
    "Calories": "kcal",
    "Total Fat, g": "g",
    "Carbs, g": "g",
    "Protein, g": "g",
    "Sodium, mg": "mg",
    "Calcium, mg": "mg",
}

# ============================================================
# BASIC HELPERS
# ============================================================
def clean_number(value):
    """Extract numeric value from strings like '1,487cals', '31g', '2,300mg', '&nbsp;g'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("&nbsp;", "").replace("&amp;nbsp;", "").replace(",", "").strip()
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def parse_day_label(value):
    """Parse day labels such as 'Saturday, May 16'. The report year is assumed to be 2026."""
    if not value:
        return None
    text = str(value).strip()
    try:
        return datetime.strptime(text + ", 2026", "%A, %b %d, %Y")
    except Exception:
        return None


def safe_filename(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", text).strip("_")


def image_to_base64(path: Path) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

# ============================================================
# EXCEL PARSING
# ============================================================
def find_header_row(ws) -> int:
    """Find nutrient header row robustly, including files with an additional Time column."""
    for r in range(1, min(ws.max_row, 40) + 1):
        values = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if "Calories" in values and any(v in values for v in ["Total Fat, g", "Carbs, g", "Protein, g"]):
            return r
    raise ValueError("Could not find nutrient header row containing Calories / macros.")


def parse_targets(ws, header_row: int) -> Dict[str, float]:
    """Read target values from rows above/including the nutrient header row."""
    targets: Dict[str, float] = {}
    for r in range(1, header_row + 1):
        for c in range(1, ws.max_column + 1):
            text = str(ws.cell(r, c).value or "")
            for nutrient, keyword in TARGET_KEYWORDS.items():
                if keyword in text:
                    value = clean_number(text)
                    if value is not None:
                        targets[nutrient] = value
    return targets


def parse_excel_file(file_path: Path) -> Tuple[Dict[str, float], pd.DataFrame]:
    """Parse one MyNetDiary-style Excel file into daily records."""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active

    header_row = find_header_row(ws)
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}
    targets = parse_targets(ws, header_row)

    records = []
    r = header_row + 1

    while r <= ws.max_row:
        date_obj = parse_day_label(ws.cell(r, 1).value)
        if date_obj is None:
            r += 1
            continue

        record = {
            "file": file_path.name,
            "person": file_path.stem,
            "date": date_obj,
            "date_label": date_obj.strftime("%b %d"),
            "main_meal_count": 0,
        }

        for nutrient in NUTRIENTS:
            col = col_map.get(nutrient)
            amount = clean_number(ws.cell(r, col).value) if col else None
            target = targets.get(nutrient)
            record[nutrient] = amount
            record[nutrient + " %target"] = amount / target * 100 if amount is not None and target else None

        # Count main meals until the next date row.
        rr = r + 1
        while rr <= ws.max_row:
            label = ws.cell(rr, 1).value
            if parse_day_label(label) is not None:
                break
            if label in MAIN_MEALS:
                cal_col = col_map.get("Calories")
                meal_cal = clean_number(ws.cell(rr, cal_col).value) if cal_col else None
                if meal_cal and meal_cal > 0:
                    record["main_meal_count"] += 1
            rr += 1

        record["included"] = record["main_meal_count"] >= MIN_MAIN_MEALS
        records.append(record)
        r = rr

    df = pd.DataFrame(records)
    if df.empty:
        raise ValueError(f"No daily records found in {file_path.name}.")
    return targets, df

# ============================================================
# STATS
# ============================================================
def robust_stats(series: pd.Series):
    """Median/IQR summary. Mean is kept as reference only."""
    s = pd.Series(series).dropna().astype(float)
    if s.empty:
        return None
    q1 = s.quantile(0.25)
    median = s.median()
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    lo = q1 - 1.5 * iqr
    hi = q3 + 1.5 * iqr
    nonout = s[(s >= lo) & (s <= hi)]
    return {
        "n": int(s.count()),
        "median": float(median),
        "q1": float(q1),
        "q3": float(q3),
        "iqr": float(iqr),
        "mean_reference": float(s.mean()),
        "min": float(s.min()),
        "max": float(s.max()),
        "lower_whisker": float(nonout.min() if len(nonout) else s.min()),
        "upper_whisker": float(nonout.max() if len(nonout) else s.max()),
        "outliers": int(((s < lo) | (s > hi)).sum()),
    }


def build_stats_df(included_df: pd.DataFrame, targets: Dict[str, float]) -> pd.DataFrame:
    rows = []
    for nutrient in NUTRIENTS:
        st = robust_stats(included_df[nutrient])
        if st is None:
            continue
        rows.append({
            "Nutrient": nutrient,
            "Unit": UNITS[nutrient],
            "Median": st["median"],
            "Q1-Q3": f"{st['q1']:.1f}-{st['q3']:.1f}",
            "Mean reference": st["mean_reference"],
            "Target": targets.get(nutrient),
            "Outliers": st["outliers"],
        })
    return pd.DataFrame(rows)

# ============================================================
# PLOTS
# ============================================================
def add_wavy_line(ax, y_frac=0.94, xcenter=1, width=0.18, amp=0.018, waves=3):
    xs, ys = [], []
    for i in range(120):
        t = i / 119
        xs.append(xcenter - width / 2 + width * t)
        ys.append(y_frac + amp * math.sin(2 * math.pi * waves * t))
    ax.plot(xs, ys, transform=ax.get_xaxis_transform(), color="black", lw=1.2, clip_on=False)


def make_boxplot(included_df: pd.DataFrame, targets: Dict[str, float], person: str, output_dir: Path) -> Path:
    fig, axes = plt.subplots(2, 3, figsize=(15, 9))
    axes = axes.flatten()

    for ax, nutrient in zip(axes, NUTRIENTS):
        vals = included_df[nutrient].dropna().astype(float)
        st = robust_stats(vals)
        unit = UNITS[nutrient]
        if vals.empty or st is None:
            ax.set_title(nutrient)
            ax.text(0.5, 0.5, "No data", ha="center", va="center")
            continue

        if nutrient == "Sodium, mg":
            target = targets.get(nutrient, 0) or 0
            upper = max(st["upper_whisker"] * 1.35, st["q3"] * 1.8, target * 1.4)
            if vals.max() > upper * 2:
                upper = min(upper, vals.max() * 0.55)
            visible = vals[vals <= upper]
            hidden = vals[vals > upper]

            ax.boxplot(vals, vert=True, showmeans=True, widths=0.45, showfliers=False)
            ax.scatter([1] * len(visible), visible, color="royalblue", alpha=0.50, s=22)
            ax.set_ylim(0, upper)
            add_wavy_line(ax)
            if len(hidden):
                ax.text(1.18, upper * 0.94, f"{len(hidden)} high outlier(s) omitted\nmax {hidden.max():.0f} mg", fontsize=8, va="top")
            ax.set_title("Sodium, mg\n(focused on median/IQR; high outliers omitted)")
        else:
            ax.boxplot(vals, vert=True, showmeans=True, widths=0.45)
            ax.scatter([1] * len(vals), vals, color="royalblue", alpha=0.50, s=22)
            ax.set_title(nutrient)

        ax.set_ylabel(unit)
        ax.set_xticks([1])
        ax.set_xticklabels(["included days"])
        ax.axhline(st["median"], color="orange", linestyle="--", linewidth=1.2)
        ax.text(1.08, st["median"], f"Median {st['median']:.0f} {unit}", va="center", fontsize=9)

        target = targets.get(nutrient)
        if target is not None:
            ax.axhline(target, color="crimson", linestyle="--", linewidth=1.3)
            y0, y1 = ax.get_ylim()
            if y0 <= target <= y1:
                ax.text(0.72, target, f"Target {target:.0f} {unit}", color="crimson", fontsize=8, va="center")

    fig.suptitle(f"{person} - Box-and-whisker by nutrient", fontsize=14)
    fig.tight_layout(rect=[0, 0, 1, 0.93])
    out = output_dir / f"{safe_filename(person)}_boxplot.png"
    fig.savefig(out, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return out


def make_combined_trend(included_df: pd.DataFrame, targets: Dict[str, float], person: str, output_dir: Path) -> Path:
    """Create one A4-friendly image with 6 nutrition trends on the same page."""
    fig, axes = plt.subplots(3, 2, figsize=(11.7, 8.3))
    axes = axes.flatten()
    x = list(included_df["date_label"])

    for i, nutrient in enumerate(TREND_NUTRIENTS):
        ax = axes[i]

        if nutrient == "Calories":
            y = included_df["Calories"]
            ax.plot(x, y, marker="o", linewidth=1.8)
            ax.set_ylabel("kcal")
            target = targets.get("Calories")
            if target is not None:
                ax.axhline(target, color="crimson", linestyle="--", linewidth=1.2, label="Target")
        else:
            pct_col = nutrient + " %target"
            y = included_df[pct_col]
            ax.plot(x, y, marker="o", linewidth=1.8)
            ax.axhline(100, color="crimson", linestyle="--", linewidth=1.2, label="Target 100%")
            ax.set_ylabel("% of target")

            vals = pd.to_numeric(y, errors="coerce").dropna()
            if len(vals) and vals.max() > 1000:
                q3 = vals.quantile(0.75)
                upper = max(300, q3 * 2.0, 150)
                ax.set_ylim(0, upper)
                add_wavy_line(ax, y_frac=0.94, xcenter=max(len(x) / 2, 1), width=max(1, len(x) * 0.08))
                high = included_df[pd.to_numeric(included_df[pct_col], errors="coerce") > upper]
                ax.text(
                    0.02, 0.95,
                    f"{len(high)} high outlier day(s) omitted\nmax {vals.max():.0f}%",
                    transform=ax.transAxes,
                    fontsize=8,
                    va="top",
                    bbox=dict(facecolor="white", alpha=0.82, edgecolor="none"),
                )
                if nutrient == "Sodium, mg":
                    for pos, (_, row) in enumerate(included_df.iterrows()):
                        pct = row[pct_col]
                        if pd.notna(pct) and pct > upper:
                            ax.annotate(f"{row['Sodium, mg']:.0f} mg", (pos, upper * 0.88), fontsize=7, rotation=25, ha="center")

        ax.set_title(nutrient)
        ax.set_xlabel("Day")
        ax.tick_params(axis="x", rotation=60, labelsize=8)
        ax.tick_params(axis="y", labelsize=8)
        ax.grid(True, alpha=0.25)
        ax.legend(loc="best", fontsize=7)

    fig.suptitle(f"{person} - Trend during record session", fontsize=15)
    fig.tight_layout(rect=[0, 0, 1, 0.94])
    out = output_dir / f"{safe_filename(person)}_combined_trend_6figures.png"
    fig.savefig(out, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return out


def make_pfc_pie(included_df: pd.DataFrame, person: str, output_dir: Path) -> Path:
    p_g = included_df["Protein, g"].dropna().mean()
    f_g = included_df["Total Fat, g"].dropna().mean()
    c_g = included_df["Carbs, g"].dropna().mean()

    values = [p_g * 4, f_g * 9, c_g * 4]
    labels = [
        f"Protein\n{p_g:.1f} g ({p_g * 4:.0f} kcal)",
        f"Fat\n{f_g:.1f} g ({f_g * 9:.0f} kcal)",
        f"Carbs\n{c_g:.1f} g ({c_g * 4:.0f} kcal)",
    ]

    fig, ax = plt.subplots(figsize=(7, 6))
    ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
    ax.axis("equal")
    ax.set_title(f"{person} - Average PFC balance")
    out = output_dir / f"{safe_filename(person)}_PFC_pie.png"
    fig.savefig(out, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return out

# ============================================================
# HTML GENERATION
# ============================================================
def format_float(v, digits=1):
    if v is None or pd.isna(v):
        return ""
    return f"{float(v):,.{digits}f}"


def df_to_html(df: pd.DataFrame) -> str:
    display_df = df.copy()
    for col in display_df.columns:
        if pd.api.types.is_numeric_dtype(display_df[col]):
            display_df[col] = display_df[col].map(lambda v: format_float(v))
    return display_df.to_html(index=False, escape=False, classes="data-table")


def create_person_html(
    person: str,
    targets: Dict[str, float],
    all_df: pd.DataFrame,
    included_df: pd.DataFrame,
    stats_df: pd.DataFrame,
    boxplot_path: Path,
    trend_path: Path,
    pie_path: Path,
    output_html: Path,
):
    summary_df = pd.DataFrame([{
        "Person/File": person,
        "Logged days": len(all_df),
        "Included days": len(included_df),
        "Excluded days": len(all_df) - len(included_df),
        "Inclusion rule": f">={MIN_MAIN_MEALS} main meals/day",
    }])

    targets_df = pd.DataFrame([{
        "Nutrient": n,
        "Target": targets.get(n),
        "Unit": UNITS[n],
    } for n in NUTRIENTS])

    boxplot_b64 = image_to_base64(boxplot_path)
    trend_b64 = image_to_base64(trend_path)
    pie_b64 = image_to_base64(pie_path)

    html_text = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{html.escape(person)} Nutrition Analysis Report</title>
<style>
@page {{ size: A4; margin: 14mm; }}
* {{ box-sizing: border-box; }}
body {{ font-family: Arial, Helvetica, sans-serif; color: #222; margin: 0; background: #f2f4f8; line-height: 1.42; }}
.report-shell {{ max-width: 900px; margin: 0 auto; background: white; }}
.page {{ width: 210mm; min-height: 297mm; padding: 14mm; margin: 0 auto 12px auto; background: white; page-break-after: always; overflow: hidden; }}
.page:last-child {{ page-break-after: auto; }}
h1 {{ color: #1f4e78; font-size: 24px; margin: 0 0 10px 0; }}
h2 {{ color: #1f4e78; font-size: 17px; border-bottom: 2px solid #1f4e78; padding-bottom: 4px; margin: 18px 0 10px 0; }}
h3 {{ color: #333; font-size: 14px; margin: 14px 0 8px 0; }}
.note {{ background: #f7f9fc; border-left: 4px solid #1f4e78; padding: 9px 11px; margin: 8px 0 12px 0; font-size: 12px; }}
.data-table {{ border-collapse: collapse; width: 100%; font-size: 11px; margin: 6px 0 12px 0; }}
.data-table th {{ background: #1f4e78; color: white; padding: 6px; text-align: left; }}
.data-table td {{ border: 1px solid #d9e2f3; padding: 6px; vertical-align: middle; }}
.data-table td:not(:first-child), .data-table th:not(:first-child) {{ text-align: center; }}
.figure {{ width: 100%; display: block; margin: 8px auto 0 auto; }}
.figure-boxplot {{ width: 100%; max-height: 146mm; object-fit: contain; }}
.figure-trend {{ width: 100%; max-height: 220mm; object-fit: contain; }}
.figure-pie {{ width: 68%; max-height: 155mm; object-fit: contain; }}
.footer-note {{ font-size: 10px; color: #666; margin-top: 8px; }}
@media screen {{ .page {{ box-shadow: 0 2px 10px rgba(0,0,0,.12); }} }}
</style>
</head>
<body>
<div class="report-shell">
  <section class="page">
    <h1>{html.escape(person)} Nutrition Analysis Report</h1>
    <div class="note">
      This report is separated by person/file. Only days with <b>>={MIN_MAIN_MEALS} main meals</b> are included. Main meals are Breakfast, Lunch, and Dinner. Snacks are not counted as main meals. Representative number = <b>median</b>; typical range = <b>Q1-Q3 / IQR</b>. Mean is shown only as reference because outliers can distort it.
    </div>

    <h2>1. Inclusion summary</h2>
    {df_to_html(summary_df)}

    <h2>2. Targets read from Excel</h2>
    {df_to_html(targets_df)}

    <h2>3. Robust nutrition summary</h2>
    {df_to_html(stats_df)}

    <h2>4. Box plot by nutrient</h2>
    <div class="note">Sodium focuses on the medium/typical range. Very high outliers are indicated with an omitted wavy line.</div>
    <img class="figure figure-boxplot" src="data:image/png;base64,{boxplot_b64}" alt="Boxplot by nutrient">
  </section>

  <section class="page">
    <h1>{html.escape(person)} - Trend during record session</h1>
    <div class="note">
      Six trend figures are shown on the same page. <b>Calories are shown in kcal</b>. Other nutrients are shown as <b>% of target</b>. Horizontal axis is day. Outlier data is considered; extreme high values may be visually omitted with a wavy line to keep the medium range readable.
    </div>
    <img class="figure figure-trend" src="data:image/png;base64,{trend_b64}" alt="Six trend figures">
  </section>

  <section class="page">
    <h1>{html.escape(person)} - PFC balance</h1>
    <div class="note">
      PFC pie chart uses included-day average grams converted to kcal: Protein = 4 kcal/g, Fat = 9 kcal/g, Carbs = 4 kcal/g.
    </div>
    <img class="figure figure-pie" src="data:image/png;base64,{pie_b64}" alt="PFC pie chart">
    <p class="footer-note">Generated from {html.escape(person)} Excel data using the inclusion rule above.</p>
  </section>
</div>
</body>
</html>
"""
    output_html.write_text(html_text, encoding="utf-8")

# ============================================================
# MAIN WORKFLOW
# ============================================================
def analyze_one_file(file_path: Path, output_dir: Path) -> Path:
    person = file_path.stem
    person_dir = output_dir / safe_filename(person)
    person_dir.mkdir(parents=True, exist_ok=True)

    targets, all_df = parse_excel_file(file_path)
    included_df = all_df[all_df["included"]].sort_values("date").copy()

    if included_df.empty:
        raise ValueError(f"No included days found in {file_path.name}. Check the meal records or MIN_MAIN_MEALS setting.")

    stats_df = build_stats_df(included_df, targets)

    # Save CSVs for QA/reproducibility.
    all_df.to_csv(person_dir / f"{safe_filename(person)}_all_daily_data.csv", index=False)
    included_df.to_csv(person_dir / f"{safe_filename(person)}_included_daily_data.csv", index=False)
    stats_df.to_csv(person_dir / f"{safe_filename(person)}_robust_stats.csv", index=False)

    boxplot_path = make_boxplot(included_df, targets, person, person_dir)
    trend_path = make_combined_trend(included_df, targets, person, person_dir)
    pie_path = make_pfc_pie(included_df, person, person_dir)

    output_html = output_dir / f"{safe_filename(person)}_nutrition_report.html"
    create_person_html(person, targets, all_df, included_df, stats_df, boxplot_path, trend_path, pie_path, output_html)
    return output_html


def resolve_files(input_dir: Path, files: List[str]) -> List[Path]:
    if files:
        paths = [Path(f) for f in files]
    else:
        paths = sorted(list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xlsm")))
    paths = [p if p.is_absolute() else input_dir / p for p in paths]
    missing = [str(p) for p in paths if not p.exists()]
    if missing:
        raise FileNotFoundError("Missing file(s): " + ", ".join(missing))
    if not paths:
        raise FileNotFoundError("No .xlsx or .xlsm files found.")
    return paths


def main(argv=None):
    parser = argparse.ArgumentParser(description="Generate separated A4 HTML nutrition reports from MyNetDiary Excel files.")
    parser.add_argument("--input-dir", default=".", help="Directory containing Excel files. Default: current folder")
    parser.add_argument("--files", nargs="*", help="Specific Excel files to process")
    parser.add_argument("--output-dir", default="nutrition_html_reports", help="Output folder")

    # parse_known_args prevents Jupyter's automatic '-f kernel.json' argument from crashing the script.
    args, _unknown = parser.parse_known_args(argv)

    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_files = resolve_files(input_dir, args.files or [])

    outputs = []
    for file_path in excel_files:
        html_path = analyze_one_file(file_path, output_dir)
        outputs.append(html_path)
        print(f"Created HTML: {html_path}")

    print("\nCompleted successfully.")
    for path in outputs:
        print(path)


if __name__ == "__main__":
    main()
